import openpyxl
from datetime import datetime
import json

file_path = 'FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Transactions']

print(f"Total rows: {ws.max_row}")
print("\nFirst 15 transactions:")
print("Row | Date | Type | Category | Description | Amount | Period")
print("-" * 120)

for row_idx in range(2, 17):
    date_val = ws.cell(row_idx, 1).value
    type_val = ws.cell(row_idx, 2).value
    category_val = ws.cell(row_idx, 3).value
    desc_val = ws.cell(row_idx, 4).value
    amount_val = ws.cell(row_idx, 5).value
    period_val = ws.cell(row_idx, 9).value
    print(f"{row_idx:3} | {str(date_val)[:10] if date_val else ''} | {str(type_val)[:10]:10} | {str(category_val)[:20]:20} | {str(desc_val)[:30]:30} | £{amount_val if amount_val else 0:8.2f} | {period_val}")

# Analyze periods and amounts
print("\n\n=== ANALYSIS BY PERIOD ===")
period_stats = {}
for row_idx in range(2, ws.max_row + 1):
    period = ws.cell(row_idx, 9).value
    amount = ws.cell(row_idx, 5).value
    type_val = ws.cell(row_idx, 2).value
    
    if period not in period_stats:
        period_stats[period] = {'count': 0, 'income': 0, 'expense': 0, 'transfer': 0}
    
    period_stats[period]['count'] += 1
    if type_val == 'Income':
        period_stats[period]['income'] += amount if amount else 0
    elif type_val == 'Expense':
        period_stats[period]['expense'] += amount if amount else 0
    elif type_val == 'Transfer':
        period_stats[period]['transfer'] += amount if amount else 0

for period in sorted([p for p in period_stats.keys() if p], key=lambda x: int(x) if str(x).isdigit() else 999):
    stats = period_stats[period]
    print(f"Period {period}: {stats['count']:3} rows | Income: £{stats['income']:8.2f} | Expense: £{stats['expense']:8.2f} | Transfer: £{stats['transfer']:8.2f}")

# Check for None/blank periods
if None in period_stats:
    print(f"\nPeriod None: {period_stats[None]['count']} rows (NOT FILTERED)")
if '' in period_stats:
    print(f"Period '': {period_stats['']['count']} rows (NOT FILTERED)")
