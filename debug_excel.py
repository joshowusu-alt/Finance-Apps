import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx', data_only=True)
ws = wb['Transactions']

# Check columns
print('First row (headers):')
for j in range(1, 10):
    print(f'  Col {j}: {ws.cell(1, j).value}')

print()
print('Looking for transactions in Period 1 date range (22 Dec 2025 - 25 Jan 2026):')

p1_start = datetime(2025, 12, 22)
p1_end = datetime(2026, 1, 25)

p1_transactions = []
for i in range(2, min(150, ws.max_row + 1)):
    date_val = ws.cell(i, 1).value
    type_val = ws.cell(i, 2).value
    cat_val = ws.cell(i, 3).value
    desc_val = ws.cell(i, 4).value
    amount_val = ws.cell(i, 5).value
    
    if not date_val or not amount_val:
        continue
        
    if isinstance(date_val, datetime):
        if p1_start <= date_val <= p1_end:
            p1_transactions.append({
                'date': date_val.strftime('%Y-%m-%d'),
                'type': type_val,
                'category': cat_val,
                'description': desc_val,
                'amount': amount_val
            })

print(f'Found {len(p1_transactions)} transactions in Period 1')
print()
print('First 15:')
for i, txn in enumerate(p1_transactions[:15]):
    print('{:<2} | {:<10} | {:<10} | {:<25} | {:<30} | {}'.format(
        i+1,
        txn['date'],
        str(txn['type'])[:10],
        str(txn['category'])[:25],
        str(txn['description'])[:30],
        txn['amount']
    ))

# Summary by category
print()
print('Category summary:')
from collections import defaultdict
cats = defaultdict(list)
for txn in p1_transactions:
    cats[txn['category']].append(txn['amount'])

for cat in sorted(cats.keys()):
    amounts = cats[cat]
    print(f'  {cat:<25} | Count: {len(amounts):>3} | Total: £{sum(amounts):>10,.2f}')
