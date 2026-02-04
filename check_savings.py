import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx', data_only=True)
ws = wb['Transactions']

p1_start = datetime(2025, 12, 22)
p1_end = datetime(2026, 1, 25)

print('All Savings Transfer transactions:')
for i in range(2, ws.max_row + 1):
    date_val = ws.cell(i, 1).value
    type_val = ws.cell(i, 2).value
    cat_val = ws.cell(i, 3).value
    desc_val = ws.cell(i, 4).value
    amount_val = ws.cell(i, 5).value
    
    if cat_val == 'Savings Transfer':
        if isinstance(date_val, datetime):
            if p1_start <= date_val <= p1_end:
                date_str = date_val.strftime('%Y-%m-%d')
                print(f'{date_str} | {type_val} | {cat_val} | {desc_val} | {amount_val}')
