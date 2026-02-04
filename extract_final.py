from openpyxl import load_workbook
import json
from datetime import datetime

excel_path = r'c:\Users\josho\OneDrive\Documents\Finance-Apps\FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx'

def map_to_type_and_category(excel_type, excel_category):
    """
    Maps Excel Type and Category to the required type and category format.
    """
    type_lower = excel_type.lower().strip() if excel_type else ""
    category_lower = (excel_category or "").lower().strip()
    
    # Map based on Type first, then Category
    if type_lower == "income":
        return "income", "income"
    
    if type_lower == "expense":
        # Map expense categories
        if "tithe" in category_lower or "giving" in category_lower:
            return "outflow", "giving"
        elif "bill" in category_lower or "electricity" in category_lower or "gas" in category_lower or "water" in category_lower or "internet" in category_lower or "phone" in category_lower:
            return "outflow", "bill"
        elif "saving" in category_lower or "investment" in category_lower:
            return "transfer", "savings"
        elif "allowance" in category_lower or "house keep" in category_lower:
            return "outflow", "allowance"
        else:
            # Default for other expenses
            return "outflow", "other"
    
    if type_lower == "transfer":
        return "transfer", "savings"
    
    # Default fallback
    return "outflow", "other"

try:
    wb = load_workbook(excel_path)
    transactions_sheet = wb['Transactions']
    
    transactions = []
    txn_id = 1
    
    for i, row in enumerate(transactions_sheet.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header row
            continue
        
        if not any(row):  # Skip empty rows
            continue
        
        # Extract columns
        date_val = row[0]
        type_val = row[1]
        category_val = row[2]
        description = row[3]
        amount = row[4]
        notes = row[7] if row[7] and not str(row[7]).startswith("=") else ""
        
        # Skip if no date or amount
        if not date_val or amount is None:
            continue
        
        # Format date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)
        
        # Map type and category
        mapped_type, mapped_category = map_to_type_and_category(type_val, category_val)
        
        # Clean amount (ensure it's numeric)
        try:
            amount_float = float(amount)
        except (ValueError, TypeError):
            continue
        
        # Create transaction
        txn = {
            "id": f"txn-{txn_id}",
            "date": date_str,
            "label": str(description) if description else "",
            "amount": amount_float,
            "type": mapped_type,
            "category": mapped_category,
            "notes": str(notes).strip() if notes else ""
        }
        
        transactions.append(txn)
        txn_id += 1
    
    # Output as JSON
    json_output = json.dumps(transactions, indent=2)
    print(json_output)
    
    # Also save to a file
    with open(r'c:\Users\josho\OneDrive\Documents\Finance-Apps\transactions_extracted.json', 'w') as f:
        f.write(json_output)
    
    print(f"\n\n// Successfully extracted {len(transactions)} transactions", file=__import__('sys').stderr)

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
