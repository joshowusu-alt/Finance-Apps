#!/usr/bin/env python3
"""
Analyze how transactions should align with budget categories.
"""
import openpyxl
from collections import defaultdict
from datetime import datetime

# Load the Excel file
wb = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx', data_only=True)

print("="*100)
print("TRANSACTIONS SHEET STRUCTURE")
print("="*100)

ws_trans = wb['Transactions']

# Get headers
headers = []
for col in range(1, 15):
    cell_value = ws_trans.cell(1, col).value
    if cell_value:
        headers.append(f"Col {col}: {cell_value}")

print("\nHeaders:")
for h in headers:
    print(f"  {h}")

# Extract Period 1 transactions (22 Dec 2025 to 25 Jan 2026)
print("\n" + "="*100)
print("PERIOD 1 TRANSACTIONS (22 Dec 2025 - 25 Jan 2026)")
print("="*100)

period1_txns = defaultdict(list)
total_income = 0
total_outflow_by_category = defaultdict(float)

for row in range(2, ws_trans.max_row + 1):
    date_val = ws_trans.cell(row, 1).value
    if not date_val:
        continue
    
    # Parse date
    if isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        try:
            date_str = str(date_val)
        except:
            continue
    
    # Check if in Period 1
    if date_str < "2025-12-22" or date_str > "2026-01-25":
        continue
    
    label = ws_trans.cell(row, 2).value or ""
    col_c = ws_trans.cell(row, 3).value or "Unknown"
    amount = ws_trans.cell(row, 4).value or 0
    
    # Parse amount
    try:
        amount = float(amount)
    except:
        continue
    
    # Determine transaction type and count
    if amount > 0 and col_c in ["Income-FM", "Income-McD", "Income-Outlier", "Income-Gifts"]:
        total_income += amount
        category = "income"
    else:
        # Map Column C to app category
        if "House Keep" in str(col_c) or "Uber" in str(col_c) or "Subscriptions" in str(col_c) or "Cosmetics" in str(col_c) or "Gifts" in str(col_c):
            app_category = "allowance"
        elif "Tithe" in str(col_c) or "Offerings" in str(col_c) or "Charity" in str(col_c) or "Donations" in str(col_c):
            app_category = "giving"
        elif "Savings" in str(col_c):
            app_category = "savings"
        elif "Income" in str(col_c):
            app_category = "income"
        else:
            app_category = "bill"
        
        total_outflow_by_category[app_category] += amount
    
    period1_txns[col_c].append({
        'date': date_str,
        'label': label,
        'amount': amount,
        'category': app_category if amount <= 0 else 'income'
    })

print(f"\nTotal Income: £{total_income:,.2f}")
print(f"\nOutflows by Category:")
for cat, total in sorted(total_outflow_by_category.items()):
    print(f"  {cat}: £{total:,.2f}")

print(f"\nTotal Outflows: £{sum(total_outflow_by_category.values()):,.2f}")
print(f"Net (Income - Outflows): £{total_income - sum(total_outflow_by_category.values()):,.2f}")

print("\n" + "="*100)
print("BREAKDOWN BY COLUMN C CATEGORY")
print("="*100)

for col_c_category in sorted(period1_txns.keys()):
    txns = period1_txns[col_c_category]
    total = sum(t['amount'] for t in txns)
    print(f"\n{col_c_category} ({len(txns)} txns, £{total:,.2f}):")
    for txn in txns[:3]:  # Show first 3
        print(f"  {txn['date']} | {txn['label'][:40]:40} | £{txn['amount']:>8.2f}")
    if len(txns) > 3:
        print(f"  ... and {len(txns)-3} more")

# Check Bills Schedule tab
print("\n\n" + "="*100)
print("BILLS SCHEDULE TAB")
print("="*100)

if 'Bills Schedule' in wb.sheetnames:
    ws_bills = wb['Bills Schedule']
    print("\nBill items:")
    for row in range(2, min(25, ws_bills.max_row + 1)):
        label = ws_bills.cell(row, 1).value
        amount = ws_bills.cell(row, 2).value
        category = ws_bills.cell(row, 3).value
        if label:
            try:
                amount = float(amount) if amount else 0
                print(f"  {label[:40]:40} | £{amount:>8.2f} | {category}")
            except:
                print(f"  {label[:40]:40} | {amount} | {category}")

# Check Budget by Period tab
print("\n\n" + "="*100)
print("BUDGET BY PERIOD TAB")
print("="*100)

if 'Budget by Period' in wb.sheetnames:
    ws_budget = wb['Budget by Period']
    print("\nFirst 15 rows of Budget by Period:")
    for row in range(1, min(16, ws_budget.max_row + 1)):
        row_data = []
        for col in range(1, 8):
            cell = ws_budget.cell(row, col)
            val = cell.value
            if val is None:
                row_data.append("")
            elif isinstance(val, (int, float)):
                row_data.append(f"{val:>10}")
            else:
                row_data.append(f"{str(val)[:20]:20}")
        print(f"Row {row}: {' | '.join(row_data)}")

wb.close()
