from openpyxl import load_workbook
import json
from datetime import datetime

excel_path = r'c:\Users\josho\OneDrive\Documents\Finance-Apps\FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx'

try:
    wb = load_workbook(excel_path)
    print("Available sheets:", wb.sheetnames)
    
    # Find transactions sheet
    transactions_sheet = None
    for sheet_name in wb.sheetnames:
        if 'transaction' in sheet_name.lower():
            transactions_sheet = wb[sheet_name]
            print(f"\nFound transactions sheet: {sheet_name}")
            break
    
    if not transactions_sheet:
        print("\nNo transactions sheet found. Checking all sheets...")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print("First 5 rows:")
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i <= 5:
                    print(f"  Row {i}: {row}")
                else:
                    break
    else:
        print("\nFirst 15 rows of transactions sheet:")
        for i, row in enumerate(transactions_sheet.iter_rows(values_only=True), 1):
            if i <= 15:
                print(f"  Row {i}: {row}")
            else:
                break
        
        # Count total rows
        total_rows = transactions_sheet.max_row
        print(f"\nTotal rows in sheet: {total_rows}")
        
        # Extract all transactions
        transactions = []
        header_row = None
        
        for i, row in enumerate(transactions_sheet.iter_rows(values_only=True), 1):
            if i == 1:
                # Assume first row is header
                header_row = row
                print(f"\nHeaders: {header_row}")
                continue
            
            if not any(row):  # Skip empty rows
                continue
            
            # Create transaction object
            txn = {
                'raw_row': i,
                'data': row
            }
            transactions.append(txn)
        
        print(f"\nFound {len(transactions)} transaction rows")
        print("\nFirst 5 transactions:")
        for txn in transactions[:5]:
            print(f"  {txn}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
