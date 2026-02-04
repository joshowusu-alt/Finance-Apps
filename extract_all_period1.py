import openpyxl
from datetime import datetime
from collections import defaultdict

# Load with data_only to get actual values
wb_values = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy.xlsx', data_only=True)
txn_sheet = wb_values['Transactions']

period1_start = datetime(2025, 12, 22)
period1_end = datetime(2026, 1, 25)

# Get all Period 1 transactions
period1_txns = []
category_counts = defaultdict(int)

print("Scanning entire sheet for Period 1 transactions...")
print(f"Sheet has {txn_sheet.max_row} rows")
print()

for row_num in range(2, txn_sheet.max_row + 1):
    date_val = txn_sheet[f'A{row_num}'].value
    
    if date_val and isinstance(date_val, datetime):
        if period1_start <= date_val <= period1_end:
            col_b_type = txn_sheet[f'B{row_num}'].value       # Type
            col_c_category = txn_sheet[f'C{row_num}'].value   # Category
            col_d_desc = txn_sheet[f'D{row_num}'].value       # Description
            col_e_amount = txn_sheet[f'E{row_num}'].value     # Amount
            
            category_counts[col_c_category] += 1
            
            period1_txns.append({
                'row': row_num,
                'date': date_val,
                'type': col_b_type,
                'category': col_c_category,
                'description': col_d_desc,
                'amount': col_e_amount
            })

# Display summary
print("="*120)
print(f"PERIOD 1 TRANSACTIONS: {len(period1_txns)} total")
print("="*120)
print()

# Group by category and show counts
print("Distribution by Column C Category:")
print("-"*120)
for cat in sorted(category_counts.keys()):
    print(f"  {cat:35s}: {category_counts[cat]:3d} transactions")

print()
print(f"Total: {len(period1_txns)} transactions")
print()

# Show first 20 transactions
print("="*120)
print("First 20 transactions (chronological):")
print("="*120)
for i, txn in enumerate(sorted(period1_txns, key=lambda x: x['date'])[:20], 1):
    print(f"{i:3d}. {txn['date'].strftime('%Y-%m-%d')} | {txn['type']:8s} | {txn['category']:30s} | "
          f"£{txn['amount']:>8} | {txn['description'][:40]}")

# Save all to JSON for processing
import json
with open('period1_full_data.json', 'w') as f:
    json.dump([{k: (v.strftime('%Y-%m-%d') if isinstance(v, datetime) else v) 
                 for k, v in txn.items() if k != 'row'} 
                for txn in sorted(period1_txns, key=lambda x: x['date'])], 
              f, indent=2)

print("\nSaved to period1_full_data.json")
