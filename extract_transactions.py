#!/usr/bin/env python
from openpyxl import load_workbook
import json
from datetime import datetime

excel_path = r'c:\Users\josho\OneDrive\Documents\Finance-Apps\FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy.xlsx'

wb = load_workbook(excel_path)
print(f"Available sheets: {wb.sheetnames}", flush=True)

# Find transactions sheet
transactions_sheet = None
for sheet_name in wb.sheetnames:
    if 'transaction' in sheet_name.lower():
        transactions_sheet = wb[sheet_name]
        print(f"Found transactions sheet: {sheet_name}", flush=True)
        break

if transactions_sheet:
    print(f"First 5 rows:", flush=True)
    for i, row in enumerate(transactions_sheet.iter_rows(values_only=True), 1):
        if i <= 5:
            print(f"Row {i}: {row}", flush=True)
        else:
            break
else:
    print("No transactions sheet found. Listing all data:", flush=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}", flush=True)
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 5:
                print(f"  Row {i}: {row}", flush=True)
            else:
                break
