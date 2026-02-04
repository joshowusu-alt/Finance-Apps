import openpyxl

excel_file = "FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx"
wb = openpyxl.load_workbook(excel_file)

print("="*120)
print("EXAMINING TRANSACTIONS SHEET - HOW ITEMS ARE CATEGORIZED")
print("="*120)

ws = wb['Transactions']

# Read headers
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print("\nHeaders:")
for i, h in enumerate(headers[:10], 1):
    print(f"  Col {i}: {h}")

print("\nSample Period 1 transactions (first 20):")
print("-"*120)

sample_count = 0
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=200, values_only=False), 2):
    date_cell = row[0].value
    type_cell = row[1].value
    cat_cell = row[2].value
    desc_cell = row[3].value
    amt_cell = row[4].value
    period_cell = row[8].value  # Column I (9th column)
    
    if period_cell == 1 and sample_count < 20:
        sample_count += 1
        # Get actual value if it's a formula
        if hasattr(period_cell, '__class__'):
            period_val = period_cell if not hasattr(period_cell, 'value') else period_cell.value
        else:
            period_val = period_cell
            
        print(f"Row {i}: Date={date_cell}, Type={type_cell}, Cat={cat_cell}, Desc={desc_cell}, Amt={amt_cell}, Period={period_val}")

# Now check Bills Schedule for the budget structure
print("\n" + "="*120)
print("BILLS SCHEDULE - INVOICE ITEMS (First 25 rows)")
print("="*120)

ws_bills = wb['Bills Schedule']
for i, row in enumerate(ws_bills.iter_rows(min_row=1, max_row=25, values_only=True), 1):
    if any(row[:8]):  # If any non-empty cell in first 8 columns
        print(f"Row {i}: {row[:8]}")

