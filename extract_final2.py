import openpyxl
import json
from datetime import datetime
from collections import defaultdict

excel_file = "FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)

# Read Bills Schedule to understand the category mapping
print("="*100)
print("BILLS SCHEDULE - DEFINES THE CATEGORY NAMES")
print("="*100)

ws_bills = wb['Bills Schedule']
bills_schedule_names = []

for i, row in enumerate(ws_bills.iter_rows(min_row=3, max_row=25, values_only=True), 1):
    bill_name = row[0]
    if bill_name and bill_name.strip() and i < 23:  # Skip the summary rows at bottom
        bills_schedule_names.append(bill_name)
        print(f"{i:2}. {bill_name}")

print(f"\nTotal bill/fixed items: {len(bills_schedule_names)}")

# Now read Transactions and group by Category column
print("\n" + "="*100)
print("TRANSACTIONS BY CATEGORY (Column C) - PERIOD 1")
print("="*100)

ws_txn = wb['Transactions']

categories_in_transactions = defaultdict(lambda: {'count': 0, 'amount': 0, 'examples': []})

for i, row in enumerate(ws_txn.iter_rows(min_row=2, max_row=1200, values_only=True), 2):
    date = row[0]
    txn_type = row[1]
    category = row[2]
    description = row[3]
    amount = row[4]
    period = row[8]  # Column I
    
    if period != 1 or not date or not amount:
        continue
    
    cat_name = str(category) if category else 'None'
    categories_in_transactions[cat_name]['count'] += 1
    categories_in_transactions[cat_name]['amount'] += abs(float(amount)) if amount else 0
    if len(categories_in_transactions[cat_name]['examples']) < 2:
        categories_in_transactions[cat_name]['examples'].append(description)

print("\nCategories in Transactions sheet (Period 1):")
for cat in sorted(categories_in_transactions.keys()):
    data = categories_in_transactions[cat]
    print(f"  '{cat:35}' Count: {data['count']:3}  Amount: £{data['amount']:8.2f}  Examples: {data['examples']}")

# Now map them to app categories based on Bills Schedule
print("\n" + "="*100)
print("MAPPING CATEGORIES TO APP CATEGORIES")
print("="*100)

# Map based on Bills Schedule
category_to_app = {
    # Bills/Fixed
    'Rent': 'bill',
    'Insurance': 'bill',
    'Road Tax': 'bill',
    'Water Bill': 'bill',
    'Electricity & Gas': 'bill',
    'Community Fibre / Internet': 'bill',
    'iPhone Payments': 'bill',
    'Parents': 'bill',
    'Credit Card Payment': 'bill',
    'Fuel': 'bill',
    'Laptop': 'bill',
    'One-Off Giving (Christmas)': 'other',
    
    # Giving
    'Tithe': 'giving',
    'Offerings': 'giving',
    'Charity - Perez Uni': 'giving',
    'Charity - JPC Utilities': 'giving',
    
    # House Keep (variable food/household)
    'House Keep': 'allowance',
    
    # Savings
    'Savings Transfer': 'savings',
}

# Extract transactions with corrected categorization
transactions = []
app_category_totals = defaultdict(lambda: {'count': 0, 'amount': 0})
unmapped = set()

for i, row in enumerate(ws_txn.iter_rows(min_row=2, max_row=1200, values_only=True), 2):
    date = row[0]
    txn_type = row[1]
    category = row[2]
    description = row[3]
    amount = row[4]
    period = row[8]  # Column I
    
    if period != 1 or not date or not amount:
        continue
    
    # Parse date
    if isinstance(date, datetime):
        date_str = date.strftime('%Y-%m-%d')
    else:
        continue
    
    # Map category to app category
    if str(category) in category_to_app:
        app_cat = category_to_app[str(category)]
    elif txn_type == 'Income' or txn_type == 'Transfer':
        if txn_type == 'Income':
            app_cat = 'income'
        else:  # Transfer
            app_cat = 'savings'
    else:
        # Unknown - map to other
        app_cat = 'other'
        if str(category) not in ['None', None, '']:
            unmapped.add(str(category))
    
    # Determine transaction type
    if txn_type == 'Transfer':
        final_type = 'transfer'
    elif txn_type == 'Income':
        final_type = 'income'
    else:
        final_type = 'outflow'
    
    label = str(description or category or '')
    
    transactions.append({
        'id': f"txn-{len(transactions)+1}",
        'date': date_str,
        'label': label,
        'amount': abs(float(amount)),
        'type': final_type,
        'category': app_cat,
        'notes': str(category) if category else '',
        'linkedRuleId': 'savings' if app_cat == 'savings' else None
    })
    
    app_category_totals[app_cat]['count'] += 1
    app_category_totals[app_cat]['amount'] += abs(float(amount))

# Sort by date
transactions.sort(key=lambda x: x['date'], reverse=True)

print("\nApp Category Totals:")
for cat in ['income', 'bill', 'giving', 'allowance', 'savings', 'other']:
    if cat in app_category_totals:
        data = app_category_totals[cat]
        print(f"  {cat:12} Count: {data['count']:3}  Amount: £{data['amount']:8.2f}")

if unmapped:
    print(f"\nUnmapped categories ({len(unmapped)}):")
    for cat in sorted(unmapped):
        count = categories_in_transactions[cat]['count']
        amt = categories_in_transactions[cat]['amount']
        print(f"  '{cat:30}' Count: {count:3}  Amount: £{amt:8.2f}")

# Save to JSON
output_file = 'period1_transactions_final.json'
with open(output_file, 'w') as f:
    json.dump(transactions, f, indent=2)

print(f"\nSaved {len(transactions)} transactions to {output_file}")

