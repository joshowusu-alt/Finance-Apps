import openpyxl
from datetime import datetime
from collections import defaultdict

wb = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx', data_only=True)
ws = wb['Transactions']

# Extract all Period 1 transactions
p1_start = datetime(2025, 12, 22)
p1_end = datetime(2026, 1, 25)

transactions_by_col_c = defaultdict(lambda: {'count': 0, 'total': 0})

for i in range(2, ws.max_row + 1):
    date_val = ws.cell(i, 1).value
    type_val = ws.cell(i, 2).value
    cat_val = ws.cell(i, 3).value  # Column C - The categorization
    desc_val = ws.cell(i, 4).value
    amount_val = ws.cell(i, 5).value
    
    if not date_val or not amount_val:
        continue
    
    if not isinstance(date_val, datetime):
        continue
        
    if not (p1_start <= date_val <= p1_end):
        continue
    
    # Skip if amount is 0
    if amount_val == 0:
        continue
    
    amount = abs(float(amount_val))
    transactions_by_col_c[cat_val]['count'] += 1
    transactions_by_col_c[cat_val]['total'] += amount

print('All transactions by Column C category:')
print()
for cat in sorted(transactions_by_col_c.keys()):
    info = transactions_by_col_c[cat]
    print('{:<35} | Count: {:>3} | Total: £{:>10,.2f}'.format(cat, info['count'], info['total']))

print()
print('Total transactions:', sum(v['count'] for v in transactions_by_col_c.values()))
