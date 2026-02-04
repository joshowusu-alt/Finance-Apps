#!/usr/bin/env python3
"""
Generate correct transactions with proper category mapping from Excel Column C (Category column).
Mapping logic based on user's budget structure:
- Income items → "income"
- House Keep, Uber - TfL, Others (subscriptions/gifts) → "allowance"
- Tithe, Offerings, Charity items, Donations → "giving"
- Savings Transfer → "savings"
- Fixed bills (Rent, Insurance, Water, Internet, Electricity, Road Tax, Fuel, Parents, Credit Card, iPhone, Laptop) → "bill"
"""
import openpyxl
import json
from datetime import datetime
from collections import defaultdict

# Define the mapping from Excel Column C (Category) to app category
CATEGORY_MAPPING = {
    # Income
    "Income - FM": "income",
    "Income - McD": "income",
    "Income - Outlier": "income",
    "Income - Gifts": "income",
    
    # Allowance (variable discretionary spending - "Others" category)
    "Others": "allowance",
    
    # Giving (tithes and charitable giving)
    "Tithe": "giving",
    "Offerings": "giving",
    "Charity - Perez Uni": "giving",
    "Charity - JPC Utilities": "giving",
    "Donations (Variable)": "giving",
    "One-Off Giving (Christmas)": "bill",  # This is one-time fixed spending, not regular giving
    
    # Savings
    "Savings Transfer": "savings",
    
    # Fixed Bills (includes House Keep which has a £560 budget)
    "Rent": "bill",
    "Insurance": "bill",
    "Road Tax": "bill",
    "Water Bill": "bill",
    "Community Fibre / Internet": "bill",
    "Electricity & Gas": "bill",
    "iPhone Payments": "bill",
    "Parents": "bill",
    "Credit Card Payment": "bill",
    "Fuel": "bill",
    "Laptop": "bill",
    "House Keep": "bill",
    "Uber - TfL": "bill",
    "Subscriptions": "bill",
    "Cosmetics": "bill",
    "Gifts": "bill",
}

# Load Excel
wb = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx', data_only=True)
ws = wb['Transactions']

# Extract Period 1 transactions
transactions = []
period1_start = datetime(2025, 12, 22)
period1_end = datetime(2026, 1, 25)

txn_id = 0
category_counts = defaultdict(lambda: {'count': 0, 'amount': 0})

for row in range(2, ws.max_row + 1):
    date_val = ws.cell(row, 1).value
    type_val = ws.cell(row, 2).value
    col_c_val = ws.cell(row, 3).value
    desc_val = ws.cell(row, 4).value
    amount_val = ws.cell(row, 5).value
    
    if not all([date_val, desc_val, col_c_val is not None, amount_val is not None]):
        continue
    
    # Check if in Period 1
    if not isinstance(date_val, datetime):
        continue
    if not (period1_start <= date_val <= period1_end):
        continue
    
    # Parse date and amount
    try:
        date_str = date_val.strftime('%Y-%m-%d')
        amount = float(amount_val)
    except:
        continue
    
    # Skip if amount is 0
    if amount == 0:
        continue
    
    # Normalize type
    type_str = str(type_val).strip().lower()
    col_c_str = str(col_c_val).strip()
    
    if type_str == 'income':
        txn_type = 'income'
        app_category = 'income'
    elif type_str == 'expense':
        txn_type = 'outflow'
        amount = abs(amount)  # Store as positive
        
        # Map Column C to category
        app_category = CATEGORY_MAPPING.get(col_c_str, 'other')
    elif type_str == 'transfer':
        txn_type = 'transfer'
        amount = abs(amount)  # Store as positive
        app_category = CATEGORY_MAPPING.get(col_c_str, 'other')
    else:
        continue
    
    txn_id += 1
    transactions.append({
        "id": f"txn-{txn_id}",
        "date": date_str,
        "label": str(desc_val),
        "amount": amount,
        "type": txn_type,
        "category": app_category,
        "notes": col_c_str,
    })
    
    # Track category stats
    category_counts[app_category]['count'] += 1
    category_counts[app_category]['amount'] += amount

# Summary
print("="*80)
print("PERIOD 1 TRANSACTION SUMMARY")
print("="*80)
print(f"Total transactions: {len(transactions)}")
print(f"\nBy category:")

for cat in ['income', 'giving', 'bill', 'allowance', 'savings', 'other']:
    if cat in category_counts:
        stats = category_counts[cat]
        count = stats['count']
        total = stats['amount']
        # Find sample txn type for this category
        sample_type = [t['type'] for t in transactions if t['category'] == cat][0]
        print(f"  {cat:15} ({count:3} txns): £{total:>10,.2f} ({sample_type})")

# Calculate totals
total_income = sum(t['amount'] for t in transactions if t['type'] == 'income')
total_outflow = sum(t['amount'] for t in transactions if t['type'] == 'outflow')
total_savings = sum(t['amount'] for t in transactions if t['type'] == 'transfer')

print(f"\n{'='*80}")
print(f"Total Income:    £{total_income:>10,.2f}")
print(f"Total Outflow:   £{total_outflow:>10,.2f}")
print(f"Total Savings:   £{total_savings:>10,.2f}")
print(f"Net (Income - Out - Savings): £{total_income - total_outflow - total_savings:>10,.2f}")

# Save transactions
with open('period1_transactions_corrected.json', 'w') as f:
    json.dump(transactions, f, indent=2)

print(f"\nSaved {len(transactions)} transactions to period1_transactions_corrected.json")

# Generate TypeScript format
print(f"\n{'='*80}")
print("Generating TypeScript format...")

ts_lines = []
for txn in transactions:
    label = str(txn["label"]).replace('"', '\\"')
    notes = str(txn["notes"]).replace('"', '\\"')
    ts_lines.append(
        f'    {{ id: "{txn["id"]}", date: "{txn["date"]}", label: "{label}", '
        f'amount: {txn["amount"]}, type: "{txn["type"]}", category: "{txn["category"]}", '
        f'notes: "{notes}", linkedRuleId: undefined }},'
    )

with open('period1_transactions_corrected.ts', 'w') as f:
    f.write('\n'.join(ts_lines))

print(f"Saved TypeScript format to period1_transactions_corrected.ts")
print(f"\nFirst 5 TypeScript lines:")
for line in ts_lines[:5]:
    print(line)

wb.close()
