import openpyxl

excel_file = "FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)

print("Available sheets:", wb.sheetnames)
print("\n" + "="*100)

# Check Bills Schedule
if 'Bills Schedule' in wb.sheetnames:
    ws = wb['Bills Schedule']
    print("\n=== BILLS SCHEDULE ===")
    print("Reading first 30 rows to understand structure...")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        if any(cell for cell in row):  # Only print non-empty rows
            print(f"Row {i}: {row}")

print("\n" + "="*100)

# Check Budget by Period
if 'Budget by Period' in wb.sheetnames:
    ws = wb['Budget by Period']
    print("\n=== BUDGET BY PERIOD ===")
    print("Reading first 40 rows...")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
        if any(cell for cell in row):  # Only print non-empty rows
            print(f"Row {i}: {row}")

print("\n" + "="*100)

# Check Category Mapping
if 'Category Mapping' in wb.sheetnames:
    ws = wb['Category Mapping']
    print("\n=== CATEGORY MAPPING ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if any(cell for cell in row):
            print(f"Row {i}: {row}")

