import openpyxl
from datetime import datetime
import json

file_path = 'FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Transactions']

transactions = []
txn_id = 1

# Map Excel categories to app categories
category_map = {
    'Income - FM': 'income',
    'Income - Outlier': 'income',
    'Income - Interest': 'income',
    'Income - Gift': 'income',
    'Electricity & Gas': 'bill',
    'Water': 'bill',
    'Internet': 'bill',
    'Mobile': 'bill',
    'Car Insurance': 'bill',
    'Tithe': 'giving',
    'Offerings': 'giving',
    'Gifts': 'giving',
    'Charity': 'giving',
    'Food': 'allowance',
    'House Keep': 'allowance',
    'Others': 'allowance',
    'Savings': 'savings',
}

# Map Excel types
type_map = {
    'Income': 'income',
    'Expense': 'outflow',
    'Transfer': 'transfer',
}

print("Extracting Period 1 transactions...")
for row_idx in range(2, ws.max_row + 1):
    period = ws.cell(row_idx, 9).value
    
    # Only include transactions where period == 1
    if period != 1:
        continue
    
    date_val = ws.cell(row_idx, 1).value
    type_val = ws.cell(row_idx, 2).value
    category_val = ws.cell(row_idx, 3).value
    desc_val = ws.cell(row_idx, 4).value
    amount_val = ws.cell(row_idx, 5).value
    notes_val = ws.cell(row_idx, 8).value
    
    # Skip if no date or amount
    if not date_val or amount_val is None:
        continue
    
    # Convert date
    if isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = str(date_val)[:10]
    
    # Map category
    app_category = category_map.get(category_val, 'other')
    
    # Map type
    app_type = type_map.get(type_val, 'outflow')
    
    # Determine transaction type based on Excel type
    if app_type == 'income':
        app_category = 'income'
    elif app_type == 'transfer':
        app_category = 'savings'
    
    # Build transaction
    txn = {
        'id': f'txn-{txn_id}',
        'date': date_str,
        'label': desc_val if desc_val else 'Transaction',
        'amount': amount_val,
        'type': app_type,
        'category': app_category,
        'notes': notes_val if notes_val else '',
    }
    
    # Auto-link transfers to savings
    if app_type == 'transfer' and app_category == 'savings':
        txn['linkedRuleId'] = 'savings'
    
    transactions.append(txn)
    txn_id += 1

print(f"Extracted {len(transactions)} Period 1 transactions")

# Summary stats
income_total = sum(t['amount'] for t in transactions if t['type'] == 'income')
expense_total = sum(t['amount'] for t in transactions if t['type'] == 'outflow')
transfer_total = sum(t['amount'] for t in transactions if t['type'] == 'transfer')

print(f"\nSummary:")
print(f"  Income: £{income_total:.2f}")
print(f"  Expenses: £{expense_total:.2f}")
print(f"  Transfers: £{transfer_total:.2f}")

# Save as JSON
with open('period1_transactions.json', 'w') as f:
    json.dump(transactions, f, indent=2)

print(f"\nSaved to period1_transactions.json")

# Also print TypeScript format
print("\n\nTypeScript array format:")
print("const transactions: Transaction[] = [")
for t in transactions[:5]:
    print(f'  {{ id: "{t["id"]}", date: "{t["date"]}", label: "{t["label"]}", amount: {t["amount"]}, type: "{t["type"]}", category: "{t["category"]}", notes: "{t["notes"]}"' + (f', linkedRuleId: "{t.get("linkedRuleId")}"' if "linkedRuleId" in t else "") + " },")
print("  // ... " + str(len(transactions) - 5) + " more transactions")
print("];")
