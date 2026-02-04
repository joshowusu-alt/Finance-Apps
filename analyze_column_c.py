import openpyxl
import json
from collections import defaultdict
from datetime import datetime

# Load Excel file with data_only=True to get calculated values
wb_formulas = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy.xlsx')
wb_values = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy.xlsx', data_only=True)

txn_sheet_formulas = wb_formulas['Transactions']
txn_sheet_values = wb_values['Transactions']

print("Excel Column Structure:")
print("="*100)
for col_num in range(1, 10):
    cell_val = txn_sheet_formulas.cell(1, col_num).value
    col_letter = chr(64 + col_num)
    print(f"Column {col_letter}: {cell_val}")

print("\n" + "="*100)
print("Looking at transaction dates to find Period 1 range (22 Dec 2025 - 25 Jan 2026)")
print("="*100)

# Find date range for Period 1
min_date = None
max_date = None
for row_num in range(2, min(50, txn_sheet_values.max_row)):
    date_val = txn_sheet_values[f'A{row_num}'].value
    if date_val and isinstance(date_val, datetime):
        if min_date is None or date_val < min_date:
            min_date = date_val
        if max_date is None or date_val > max_date:
            max_date = date_val

print(f"Date range observed: {min_date} to {max_date}")
print(f"Period 1 should be: 22 Dec 2025 to 25 Jan 2026")

# Count transactions by Column C Category where date is in Period 1
print("\n" + "="*100)
print("Period 1 Transactions (by Column C Category):")
print("="*100)

period1_start = datetime(2025, 12, 22)
period1_end = datetime(2026, 1, 25)

category_counts = defaultdict(int)
category_samples = defaultdict(list)
all_period1_txns = []

for row_num in range(2, txn_sheet_values.max_row + 1):
    date_val = txn_sheet_values[f'A{row_num}'].value
    col_c_category = txn_sheet_values[f'C{row_num}'].value  # Column C = Category
    col_d_description = txn_sheet_values[f'D{row_num}'].value  # Column D = Description
    col_e_amount = txn_sheet_values[f'E{row_num}'].value  # Column E = Amount
    
    # Check if date is in Period 1
    if date_val and isinstance(date_val, datetime):
        if period1_start <= date_val <= period1_end:
            category_counts[col_c_category] += 1
            all_period1_txns.append({
                'date': date_val,
                'category': col_c_category,
                'description': col_d_description,
                'amount': col_e_amount
            })
            
            if len(category_samples[col_c_category]) < 2:
                category_samples[col_c_category].append((col_d_description, col_e_amount))

print("\nTransaction count by Column C Category:")
for cat in sorted(category_counts.keys()):
    count = category_counts[cat]
    print(f"\n{cat}: {count} transactions")
    for desc, amt in category_samples[cat]:
        print(f"  - {str(desc)[:60]:60s} £{amt}")

print("\n" + "="*100)
print(f"Total Period 1 transactions: {len(all_period1_txns)}")
print("="*100)

# Show the mapping: Column C category -> App category
print("\nColumn C Category Mapping to App Categories:")
print("="*100)
mapping = {
    'Income - FM': 'income',
    'Income - McD': 'income',
    'Income - Outlier': 'income',
    'Income - Gifts': 'income',
    'House Keep': 'bill',
    'Electricity & Gas': 'bill',
    'Water Bill': 'bill',
    'Rent': 'bill',
    'Insurance': 'bill',
    'Road Tax': 'bill',
    'Internet': 'bill',
    'iPhone Payments': 'bill',
    'Parents': 'bill',
    'Credit Card Payment': 'bill',
    'Fuel': 'bill',
    'Laptop': 'bill',
    'One-Off Giving (Christmas)': 'bill',
    'Tithe': 'giving',
    'Offerings': 'giving',
    'Charity - Perez Uni': 'giving',
    'Charity - JPC Utilities': 'giving',
    'Donations (Variable)': 'giving',
    'Others': 'allowance',
    'Uber - TfL': 'allowance',
    'Food': 'bill',
    'Savings Transfer': 'savings',
}

for col_c_cat in sorted(category_counts.keys()):
    app_cat = mapping.get(col_c_cat, '???')
    print(f"{col_c_cat:30s} -> {app_cat}")
