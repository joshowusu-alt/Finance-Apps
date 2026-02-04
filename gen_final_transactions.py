#!/usr/bin/env python3
"""
Generate correct transactions mapped to app categories.
Uses Budget by Period tab structure as source of truth.
"""
import openpyxl
import json
from datetime import datetime
from collections import defaultdict

# Correct mapping based on Budget by Period tab
CATEGORY_MAPPING = {
    # INCOME
    "Income - FM": "income",
    "Income - McD": "income",
    "Income - Outlier": "income",
    "Income - Gifts": "income",
    
    # GIVING
    "Tithe": "giving",
    "Offerings": "giving",
    "Charity - Perez Uni": "giving",
    "Charity - JPC Utilities": "giving",
    "Donations (Variable)": "giving",
    
    # FIXED (all bills including One-Off Giving)
    "Rent": "bill",
    "Insurance": "bill",
    "Road Tax": "bill",
    "Water Bill": "bill",
    "Community Fibre / Internet": "bill",
    "Electricity & Gas": "bill",
    "iPhone Payments": "bill",
    "Parents": "bill",
    "Credit Card Payment": "bill",
    "Fuel": "bill",
    "House Keep": "bill",
    "Laptop": "bill",
    "One-Off Giving (Christmas)": "bill",
    
    # VARIABLE
    "Food": "bill",
    "Others": "bill",
    
    # SAVINGS
    "Savings Transfer": "savings",
}

# Load Excel
wb = openpyxl.load_workbook('FINAL_2026_Cashflow_System_AppleStyle_v3_Start22Dec2025 - Copy (2).xlsx', data_only=True)
ws = wb['Transactions']

# Extract Period 1 transactions
transactions = []
period1_start = datetime(2025, 12, 22)
period1_end = datetime(2026, 1, 25)

txn_id = 0
category_totals = defaultdict(float)

for row in range(2, ws.max_row + 1):
    date_val = ws.cell(row, 1).value
    type_val = ws.cell(row, 2).value  # "Expense", "Income", "Transfer"
    col_c_val = ws.cell(row, 3).value  # Column C - The category
    label_val = ws.cell(row, 4).value
    amount_val = ws.cell(row, 5).value
    
    if not all([date_val, col_c_val, amount_val]):
        continue
    
    if not isinstance(date_val, datetime):
        continue
    
    # Check if in Period 1
    if not (period1_start <= date_val <= period1_end):
        continue
    
    # Skip zero amounts
    if amount_val == 0:
        continue
    
    # Parse amount
    try:
        amount = float(amount_val)
    except:
        continue
    
    date_str = date_val.strftime('%Y-%m-%d')
    col_c_str = str(col_c_val).strip()
    
    # Determine type and category
    if amount > 0:
        txn_type = "income"
        category = "income"
    elif type_val and "Transfer" in str(type_val):
        txn_type = "transfer"
        amount = abs(amount)
        category = CATEGORY_MAPPING.get(col_c_str, "savings")
    else:
        txn_type = "outflow"
        amount = abs(amount)
        category = CATEGORY_MAPPING.get(col_c_str, "bill")
    
    txn_id += 1
    transactions.append({
        "id": f"txn-{txn_id}",
        "date": date_str,
        "label": label_val,
        "amount": amount,
        "type": txn_type,
        "category": category,
        "notes": col_c_str,
    })
    
    category_totals[category] += amount

# Summary
print("="*80)
print("PERIOD 1 TRANSACTION SUMMARY (Corrected Mapping)")
print("="*80)
print(f"Total transactions: {len(transactions)}")
print(f"\nBy category:")
for cat in ["income", "giving", "bill", "savings"]:
    txns = [t for t in transactions if t['category'] == cat]
    if txns:
        total = sum(t['amount'] for t in txns)
        txn_type = txns[0]['type']
        print(f"  {cat:15} ({len(txns):3} txns): £{total:>10,.2f} ({txn_type})")

# Calculate totals
total_income = sum(t['amount'] for t in transactions if t['type'] == 'income')
total_giving = sum(t['amount'] for t in transactions if t['category'] == 'giving')
total_bills = sum(t['amount'] for t in transactions if t['category'] == 'bill')
total_savings = sum(t['amount'] for t in transactions if t['type'] == 'transfer')

print(f"\n{'='*80}")
print(f"Income:       £{total_income:>10,.2f}")
print(f"Giving:       £{total_giving:>10,.2f}")
print(f"Bills:        £{total_bills:>10,.2f}")
print(f"Savings:      £{total_savings:>10,.2f}")
print(f"Total Out:    £{total_giving + total_bills + total_savings:>10,.2f}")

# Save transactions
with open('period1_transactions_final.json', 'w') as f:
    json.dump(transactions, f, indent=2)

print(f"\nSaved {len(transactions)} transactions to period1_transactions_final.json")

# Generate TypeScript format
print(f"\n{'='*80}")
print("Generating TypeScript format...")

ts_lines = []
for txn in transactions:
    label = txn["label"].replace('"', '\\"')
    notes = txn["notes"].replace('"', '\\"')
    ts_lines.append(
        f'    {{ id: "{txn["id"]}", date: "{txn["date"]}", label: "{label}", '
        f'amount: {txn["amount"]}, type: "{txn["type"]}", category: "{txn["category"]}", '
        f'notes: "{notes}", linkedRuleId: undefined }},'
    )

with open('period1_transactions_final.ts', 'w') as f:
    f.write('\n'.join(ts_lines))

print(f"Saved TypeScript format to period1_transactions_final.ts")

wb.close()
